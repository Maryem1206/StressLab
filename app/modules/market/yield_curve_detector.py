"""
yield_curve_detector.py
=======================
Auto-detection of the yield curve Excel file among session uploads.

Two-level detection
-------------------
Level 1 -- Filename scoring (fast).
  Scan all .xlsx/.xls files in the upload directory, score each by keyword
  presence, and retain the highest-scoring file if at least one high-priority
  keyword matched.

Level 2 -- Content inspection (fallback when L1 is ambiguous).
  Open each candidate with openpyxl and verify it contains a sheet whose
  name contains a yield-curve keyword AND enough maturity-column headers.

Override
--------
MarketRiskModule(excel_path=...) bypasses the detector entirely.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import List, Optional, Tuple

LOG = logging.getLogger("market.yield_curve_detector")

# -- Keyword scoring table ---------------------------------------------------
YIELD_CURVE_KEYWORDS = {
    "haute_priorite"  : ["yield", "curve", "yieldcurve", "taux", "courbe",
                         "tbill", "tbond", "sovereign"],
    "priorite_moyenne": ["rate", "bond", "treasury", "interest", "egypt",
                         "egp", "cbe"],
    "priorite_basse"  : ["historical", "data", "market", "stresslab", "template"],
}
_SCORE_WEIGHTS = {"haute_priorite": 10, "priorite_moyenne": 3, "priorite_basse": 1}

# -- Content validation constants -------------------------------------------
_YIELD_SHEET_KW  = {"yield", "curve", "taux", "rate"}
_YIELD_COL_KW    = {"T91", "T182", "T273", "T364", "T3Y", "T5Y", "T10Y", "date"}
_MIN_COL_MATCHES = 3


class AmbiguousFileError(Exception):
    """Raised when several candidate yield-curve files are detected."""

    def __init__(self, candidates: List[Path]) -> None:
        self.candidates = candidates
        super().__init__(
            f"Multiple yield-curve file candidates: "
            f"{[str(c) for c in candidates]}. "
            "Specify excel_path explicitly to disambiguate."
        )


# ---------------------------------------------------------------------------
# Level-1 helpers
# ---------------------------------------------------------------------------

def _normalise_stem(stem: str) -> str:
    return re.sub(r"[-_\s]+", " ", stem.lower())


def _score_filename(path: Path) -> Tuple[int, bool]:
    """Return (score, has_high_priority_keyword)."""
    stem  = _normalise_stem(path.stem)
    score = 0
    has_high = False
    for tier, keywords in YIELD_CURVE_KEYWORDS.items():
        w = _SCORE_WEIGHTS[tier]
        for kw in keywords:
            if kw in stem:
                score += w
                if tier == "haute_priorite":
                    has_high = True
    return score, has_high


# ---------------------------------------------------------------------------
# Level-2 helpers
# ---------------------------------------------------------------------------

def _validate_by_content(path: Path) -> bool:
    """Return True if the file contains a yield-curve sheet + maturity columns."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for shname in wb.sheetnames:
            if any(kw in shname.lower() for kw in _YIELD_SHEET_KW):
                ws = wb[shname]
                headers: List[str] = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = [str(c) if c is not None else "" for c in row]
                    break
                matches = sum(
                    1 for kw in _YIELD_COL_KW
                    if any(kw.lower() in h.lower() for h in headers)
                )
                if matches >= _MIN_COL_MATCHES:
                    wb.close()
                    return True
        wb.close()
    except Exception as exc:
        LOG.debug("Content inspection failed for %s: %s", path, exc)
    return False


# ---------------------------------------------------------------------------
# Main detector class
# ---------------------------------------------------------------------------

class YieldCurveFileDetector:
    """Detect the yield curve Excel file in an upload directory.

    Parameters
    ----------
    upload_dir : str or Path
        Directory that contains the user-uploaded files for the current session.

    Methods
    -------
    detect() -> Path
        Return the single best matching file or raise:
          - AmbiguousFileError  if multiple candidates score equally
          - FileNotFoundError   if no valid candidate is found

    Usage
    -----
    path = YieldCurveFileDetector("/uploads/session_123").detect()
    """

    def __init__(self, upload_dir: str | Path) -> None:
        self.upload_dir = Path(upload_dir)

    def detect(self) -> Path:
        """Return the single best matching yield-curve file."""
        xlsx_files = (
            list(self.upload_dir.glob("*.xlsx"))
            + list(self.upload_dir.glob("*.xls"))
        )
        if not xlsx_files:
            raise FileNotFoundError(
                f"No .xlsx / .xls files in upload directory: {self.upload_dir}. "
                "Please upload a yield curve file."
            )

        # ---- Level 1: filename scoring ----
        scored = [(f, *_score_filename(f)) for f in xlsx_files]
        high   = [(f, sc) for f, sc, has_high in scored if has_high]

        if len(high) == 1:
            LOG.info("Yield curve file detected (L1-unique): %s", high[0][0].name)
            return high[0][0]

        if len(high) > 1:
            high.sort(key=lambda x: x[1], reverse=True)
            if high[0][1] > high[1][1]:
                LOG.info("Yield curve file detected (L1-score): %s", high[0][0].name)
                return high[0][0]
            l2_candidates = [f for f, _ in high]
        else:
            l2_candidates = xlsx_files

        # ---- Level 2: content inspection ----
        LOG.info(
            "L1 ambiguous -- inspecting content of %d file(s)", len(l2_candidates)
        )
        l2_ok = [f for f in l2_candidates if _validate_by_content(f)]

        if len(l2_ok) == 1:
            LOG.info("Yield curve file detected (L2): %s", l2_ok[0].name)
            return l2_ok[0]
        if len(l2_ok) > 1:
            raise AmbiguousFileError(l2_ok)

        raise FileNotFoundError(
            f"No yield-curve file detected in {self.upload_dir}. "
            "Expected a file with a sheet named 'yield'/'curve'/'taux'/'rate' "
            "and columns T91j/T182j/T273j/T364j/T3Y/T5Y/T10Y. "
            "Upload the correct file or pass excel_path explicitly."
        )